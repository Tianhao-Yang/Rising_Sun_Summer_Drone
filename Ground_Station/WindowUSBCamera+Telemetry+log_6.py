import cv2
import numpy as np
import time
import threading
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import math
from collections import deque

import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment
from pymavlink import mavutil


# =========================
# USB camera settings
# =========================
CAMERA_INDEX = 1
WINDOW_NAME = "USB Camera"
CAMERA_RECONNECT_INTERVAL = 2.0


# =========================
# MAVLink telemetry settings
# =========================
TELEMETRY_PORT = "COM8"
TELEMETRY_BAUD = 57600

TELEMETRY_RECONNECT_INTERVAL = 3.0
HEARTBEAT_TIMEOUT = 3.0
RC_SIGNAL_TIMEOUT = 2.0

# Re-send MAVLink message-rate requests periodically after connection.
# This fixes the startup-order case where Python starts before the Pixhawk/
# telemetry radio has fully finished booting.
MESSAGE_REQUEST_RETRY_INTERVAL = 3.0

# Your four motors are connected to outputs 9, 10, 11 and 12.
MOTOR_OUTPUT_CHANNELS = (9, 10, 11, 12)

# ArduPilot reports commanded motor output on a 1000-2000 scale.
OUTPUT_MIN = 1000
OUTPUT_MAX = 2000


# =========================
# Logging settings
# =========================
LOG_SAMPLE_INTERVAL = 0.1  # 10 samples per second

# This assumes the script is inside:
# Code/HUD/Control.py
# and logs should be stored inside:
# Code/Log/
BASE_LOG_DIRECTORY = (
    Path(__file__).resolve().parent.parent / "Log"
)


# =========================
# HUD layout settings
# =========================

# Empty border around the normal HUD.
# Normal green HUD graphics will not be copied into this margin.
HUD_SAFE_MARGIN = 25

# Compass may extend below the bottom of the screen.
# 0   = compass bottom touches the screen bottom
# 20  = compass extends 20 px below the screen
# 40  = compass extends 40 px below the screen
COMPASS_DOWN_OFFSET = 10


# =========================
# HUD alert thresholds
# =========================
# TEL / RC:
#   GREEN  >= warning threshold
#   YELLOW between warning and critical
#   RED    < critical
TEL_WARNING_PERCENT = 40
TEL_CRITICAL_PERCENT = 20

RC_WARNING_PERCENT = 40
RC_CRITICAL_PERCENT = 20

# 4S LiPo total voltage:
#   GREEN  >= 14.0 V
#   YELLOW 13.2 V .. <14.0 V
#   RED    < 13.2 V
BAT_VOLTAGE_WARNING = 14.0
BAT_VOLTAGE_CRITICAL = 13.2

# Total aircraft current:
#   GREEN  <= 60 A
#   YELLOW >60 A .. <=90 A
#   RED    >90 A
CURRENT_WARNING_A = 60.0
CURRENT_CRITICAL_A = 90.0

# GPS is only GREEN / RED.
# GREEN requires all three:
#   fix_type >= 3
#   satellites > 3
#   HDOP <= 2.0
GPS_MIN_FIX_TYPE = 3
GPS_MIN_SATELLITES_EXCLUSIVE = 3
GPS_MAX_HDOP = 2.0

# Critical warning text blink period.
WARNING_BLINK_PERIOD_S = 1.0

# Manual HUD alert test:
#   Ctrl + 0 = normal live thresholds / reset
#   Ctrl + 1 = force TEL / RC / CUR / BAT into YELLOW warning state
#   Ctrl + 2 = force TEL / RC / CUR / BAT into RED critical state
#
# The OpenCV HUD window must be active.


@dataclass
class TelemetryState:
    connected: bool = False
    status: str = "Telemetry disconnected"

    battery_voltage_v: float | None = None
    battery_remaining_percent: int | None = None
    total_current_a: float | None = None

    # Link-quality readouts for the lower-left status panel.
    #
    # telemetry_link_quality_percent is calculated locally from MAVLink
    # packet sequence continuity, similar in meaning to Mission Planner's
    # telemetry connection quality percentage.
    telemetry_link_quality_percent: int | None = None
    rc_rssi_percent: int | None = None

    # Main flight HUD values.
    # altitude_m is relative to the home/arming position when available.
    altitude_m: float | None = None
    ground_speed_m_s: float | None = None
    vertical_speed_m_s: float | None = None
    heading_deg: float | None = None

    # GPS status from GPS_RAW_INT.
    # MAVLink fix_type >= 3 means a valid 3D GPS fix.
    gps_fix_type: int | None = None
    gps_satellites_visible: int | None = None
    gps_hdop: float | None = None
    pitch_deg: float | None = None
    roll_deg: float | None = None
    yaw_deg: float | None = None

    motor_percentages: list[float | None] = field(
        default_factory=lambda: [None, None, None, None]
    )

    # Safety-switch/logging state.
    motor_outputs_enabled: bool | None = None
    recording: bool = False
    recording_start_time: float | None = None
    records: list[dict] = field(default_factory=list)
    last_log_sample_time: float = 0.0

    last_heartbeat_time: float = 0.0
    last_rc_message_time: float = 0.0

    # RC receiver state.
    # True  = ArduPilot reports Radio/RC failsafe
    # False = RC receiver is healthy
    # None  = not determined yet
    rc_failsafe: bool | None = None

    lock: threading.Lock = field(default_factory=threading.Lock, repr=False)


def create_no_camera_screen():
    # Keep the fallback black frame when the USB camera is disconnected,
    # but do not draw any "No USB Camera Detected" text.
    screen = np.zeros((480, 640, 3), dtype=np.uint8)
    return screen


def open_usb_camera():
    cap = cv2.VideoCapture(CAMERA_INDEX, cv2.CAP_DSHOW)

    if not cap.isOpened():
        cap.release()
        return None

    return cap


def pwm_equivalent_to_percent(output_value):
    """
    Convert ArduPilot's 1000-2000 motor-output representation to 0-100%.

    This is the commanded output percentage, not actual mechanical RPM.
    """
    if output_value is None or output_value == 0:
        return None

    percentage = (
        (output_value - OUTPUT_MIN)
        / (OUTPUT_MAX - OUTPUT_MIN)
        * 100.0
    )
    

    return max(0.0, min(100.0, percentage))


def request_message_interval(master, message_id, frequency_hz):
    """
    Ask ArduPilot to send a MAVLink message at the requested frequency.
    """
    interval_us = int(1_000_000 / frequency_hz)

    master.mav.command_long_send(
        master.target_system,
        master.target_component,
        mavutil.mavlink.MAV_CMD_SET_MESSAGE_INTERVAL,
        0,
        message_id,
        interval_us,
        0,
        0,
        0,
        0,
        0,
    )


def request_required_messages(master):
    # SYS_STATUS:
    # battery voltage/current and safety-switch state.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_SYS_STATUS,
        10,
    )

    # BATTERY_STATUS: another source for battery voltage/current.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_BATTERY_STATUS,
        5,
    )

    # SERVO_OUTPUT_RAW: motor output commands for channels 9-12.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_SERVO_OUTPUT_RAW,
        10,
    )

    # VFR_HUD: ground speed and heading.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_VFR_HUD,
        10,
    )

    # GLOBAL_POSITION_INT: altitude relative to home in millimetres.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_GLOBAL_POSITION_INT,
        10,
    )

    # GPS_RAW_INT: GPS receiver fix status.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_GPS_RAW_INT,
        5,
    )

    # ATTITUDE: pitch/roll/yaw. We use pitch for the HUD pitch ladder.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_ATTITUDE,
        20,
    )

    # RC_CHANNELS: receiver RSSI.
    request_message_interval(
        master,
        mavutil.mavlink.MAVLINK_MSG_ID_RC_CHANNELS,
        5,
    )



def close_master(master):
    if master is not None:
        try:
            master.close()
        except Exception:
            pass


def automatic_axis_max(values, minimum_max):
    """
    Calculate a readable automatic y-axis maximum with some headroom.
    """
    valid_values = []

    for value in values:
        if value is None:
            continue

        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            continue

        if math.isfinite(numeric_value):
            valid_values.append(numeric_value)

    if not valid_values:
        return minimum_max

    data_max = max(valid_values)

    if data_max <= 0:
        return minimum_max

    target = max(data_max * 1.10, minimum_max)

    exponent = math.floor(math.log10(target))
    magnitude = 10 ** exponent
    normalized = target / magnitude

    if normalized <= 1:
        nice_normalized = 1
    elif normalized <= 2:
        nice_normalized = 2
    elif normalized <= 5:
        nice_normalized = 5
    else:
        nice_normalized = 10

    return nice_normalized * magnitude


def save_recording(records):
    """
    Save one recording session to Excel and PNG files.

    All filenames use Control instead of MissionPlanner.
    """
    if not records:
        print("No log data was collected. Nothing was saved.")
        return

    session_name = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    session_directory = BASE_LOG_DIRECTORY / session_name
    session_directory.mkdir(parents=True, exist_ok=True)

    excel_path = session_directory / "Control_Log.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Control Log"

    headers = [
        "Elapsed Time (s)",
        "Computer Time",
        "Battery Current (A)",
        "Motor 1 (%)",
        "Motor 2 (%)",
        "Motor 3 (%)",
        "Motor 4 (%)",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for record in records:
        worksheet.append(
            [
                record["elapsed_time_s"],
                record["computer_time"],
                record["current_amps"],
                record["motor_1_percent"],
                record["motor_2_percent"],
                record["motor_3_percent"],
                record["motor_4_percent"],
            ]
        )

    column_widths = {
        "A": 18,
        "B": 24,
        "C": 20,
        "D": 15,
        "E": 15,
        "F": 15,
        "G": 15,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=2):
        row[0].number_format = "0.000"
        row[2].number_format = "0.00"

        for cell in row[3:7]:
            cell.number_format = "0.00"

    # Excel motor chart.
    motor_chart = LineChart()
    motor_chart.title = "Motor Output Commands"
    motor_chart.y_axis.title = "Motor Output (%)"
    motor_chart.x_axis.title = "Elapsed Time (s)"
    motor_chart.height = 10
    motor_chart.width = 20

    motor_data = Reference(
        worksheet,
        min_col=4,
        max_col=7,
        min_row=1,
        max_row=worksheet.max_row,
    )

    time_categories = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=worksheet.max_row,
    )

    motor_chart.add_data(motor_data, titles_from_data=True)
    motor_chart.set_categories(time_categories)
    worksheet.add_chart(motor_chart, "I2")

    # Excel current chart.
    current_chart = LineChart()
    current_chart.title = "Battery Current"
    current_chart.y_axis.title = "Current (A)"
    current_chart.x_axis.title = "Elapsed Time (s)"
    current_chart.height = 10
    current_chart.width = 20

    current_data = Reference(
        worksheet,
        min_col=3,
        min_row=1,
        max_row=worksheet.max_row,
    )

    current_chart.add_data(current_data, titles_from_data=True)
    current_chart.set_categories(time_categories)
    worksheet.add_chart(current_chart, "I22")

    workbook.save(excel_path)

    elapsed_times = [
        record["elapsed_time_s"]
        for record in records
    ]

    current_values = [
        record["current_amps"]
        for record in records
    ]

    motor_values = [
        [
            record[f"motor_{motor_number}_percent"]
            for record in records
        ]
        for motor_number in range(1, 5)
    ]

    # Individual motor plots.
    for motor_number, values in enumerate(motor_values, start=1):
        y_axis_max = automatic_axis_max(values, minimum_max=10)

        plt.figure(figsize=(10, 5))
        plt.plot(elapsed_times, values)
        plt.title(f"Motor {motor_number} Output")
        plt.xlabel("Elapsed Time (s)")
        plt.ylabel("Motor Output (%)")
        plt.ylim(0, y_axis_max)
        plt.grid(True)
        plt.tight_layout()

        output_path = (
            session_directory
            / f"Control_Log_Motor_{motor_number}.png"
        )

        plt.savefig(output_path, dpi=150)
        plt.close()

    # Combined motor plot.
    all_motor_values = [
        value
        for motor_data_list in motor_values
        for value in motor_data_list
    ]

    all_motor_y_max = automatic_axis_max(
        all_motor_values,
        minimum_max=10,
    )

    plt.figure(figsize=(11, 6))

    for motor_number, values in enumerate(motor_values, start=1):
        plt.plot(
            elapsed_times,
            values,
            label=f"Motor {motor_number}",
        )

    plt.title("All Motor Output Commands")
    plt.xlabel("Elapsed Time (s)")
    plt.ylabel("Motor Output (%)")
    plt.ylim(0, all_motor_y_max)
    plt.grid(True)
    plt.legend()
    plt.tight_layout()

    all_motors_path = (
        session_directory
        / "Control_Log_All_Motors.png"
    )

    plt.savefig(all_motors_path, dpi=150)
    plt.close()

    # Battery-current plot.
    current_y_max = automatic_axis_max(
        current_values,
        minimum_max=5,
    )

    plt.figure(figsize=(10, 5))
    plt.plot(elapsed_times, current_values)
    plt.title("Battery Current")
    plt.xlabel("Elapsed Time (s)")
    plt.ylabel("Current (A)")
    plt.ylim(0, current_y_max)
    plt.grid(True)
    plt.tight_layout()

    current_path = (
        session_directory
        / "Control_Log_Battery_Current.png"
    )

    plt.savefig(current_path, dpi=150)
    plt.close()

    print(f"Control log saved to: {session_directory}")
    print(f"Excel file: {excel_path.name}")


def start_recording(state, now):
    with state.lock:
        state.recording = True
        state.recording_start_time = now
        state.records = []
        state.last_log_sample_time = 0.0

    print("Safety released: Control logging started.")


def stop_recording(state):
    with state.lock:
        state.recording = False
        records_to_save = list(state.records)
        state.records = []
        state.recording_start_time = None

    print("Safety enabled: Control logging stopped.")

    # Save in a separate thread so telemetry reception does not pause.
    save_thread = threading.Thread(
        target=save_recording,
        args=(records_to_save,),
        daemon=True,
    )

    save_thread.start()


def process_safety_state(state, outputs_enabled, now):
    """
    Start logging when the safety switch enables motor outputs.
    Stop and save when the safety switch disables motor outputs again.

    The first received state only initializes the detector. This prevents the
    program from unexpectedly starting a log if it was launched after Safety
    had already been released.
    """
    with state.lock:
        previous_state = state.motor_outputs_enabled
        state.motor_outputs_enabled = outputs_enabled

    if previous_state is None:
        initial_text = (
            "OUTPUTS ENABLED"
            if outputs_enabled
            else "SAFETY ON"
        )
        print(f"Initial safety state: {initial_text}")
        return

    if not previous_state and outputs_enabled:
        start_recording(state, now)

    elif previous_state and not outputs_enabled:
        with state.lock:
            was_recording = state.recording

        if was_recording:
            stop_recording(state)


def append_log_sample_if_needed(state, now):
    with state.lock:
        if not state.recording:
            return

        if state.recording_start_time is None:
            return

        if (
            now - state.last_log_sample_time
            < LOG_SAMPLE_INTERVAL
        ):
            return

        state.last_log_sample_time = now

        motor_values = list(state.motor_percentages)

        record = {
            "elapsed_time_s": now - state.recording_start_time,
            "computer_time": datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S.%f"
            )[:-3],
            "current_amps": state.total_current_a,
            "motor_1_percent": motor_values[0],
            "motor_2_percent": motor_values[1],
            "motor_3_percent": motor_values[2],
            "motor_4_percent": motor_values[3],
        }

        state.records.append(record)




def update_mavlink_packet_quality(
    msg,
    now,
    last_seq_by_source,
    quality_events,
    window_seconds=5.0,
):
    """
    Estimate telemetry link quality from MAVLink packet sequence continuity.

    MAVLink packets from each source contain an 8-bit sequence number.
    If the sequence jumps forward, the missing sequence numbers are counted
    as lost packets.

    A rolling time window is used so the displayed percentage responds to
    recent link quality instead of being an all-time average.

    Returns:
        int 0..100 when enough information is available, otherwise None.
    """
    try:
        seq = int(msg.get_seq()) & 0xFF
        src_system = int(msg.get_srcSystem())
        src_component = int(msg.get_srcComponent())
    except Exception:
        return None

    source_key = (
        src_system,
        src_component,
    )

    previous_seq = last_seq_by_source.get(source_key)
    last_seq_by_source[source_key] = seq

    # First packet from a source establishes the baseline.
    if previous_seq is None:
        quality_events.append(
            (
                now,
                1,  # received
                0,  # lost
            )
        )
    else:
        delta = (seq - previous_seq) & 0xFF

        # delta == 0 can be a duplicate packet.
        # Very large deltas usually indicate a reset/out-of-order stream rather
        # than hundreds of genuinely lost packets, so re-baseline instead.
        if delta == 0:
            pass
        elif 1 <= delta <= 127:
            lost_packets = max(
                0,
                delta - 1,
            )

            quality_events.append(
                (
                    now,
                    1,
                    lost_packets,
                )
            )
        else:
            # Sequence moved backwards/reset; do not treat it as massive loss.
            quality_events.append(
                (
                    now,
                    1,
                    0,
                )
            )

    cutoff = now - window_seconds

    while (
        quality_events
        and quality_events[0][0] < cutoff
    ):
        quality_events.popleft()

    received_total = sum(
        event[1]
        for event in quality_events
    )

    lost_total = sum(
        event[2]
        for event in quality_events
    )

    expected_total = (
        received_total
        + lost_total
    )

    if expected_total <= 0:
        return None

    quality = (
        received_total
        / expected_total
        * 100.0
    )

    return int(round(
        max(
            0.0,
            min(
                100.0,
                quality,
            ),
        )
    ))

def telemetry_worker(state, stop_event):
    master = None
    last_message_request_time = 0.0

    # Packet-loss tracker used for the TEL percentage.
    # Keep sequence history separately for each MAVLink source component.
    last_seq_by_source = {}
    quality_events = deque()

    while not stop_event.is_set():
        if master is None:
            with state.lock:
                state.connected = False
                state.status = f"Connecting to {TELEMETRY_PORT}..."
                state.telemetry_link_quality_percent = None

            last_seq_by_source.clear()
            quality_events.clear()

            try:
                master = mavutil.mavlink_connection(
                    TELEMETRY_PORT,
                    baud=TELEMETRY_BAUD,
                    autoreconnect=True,
                    source_system=255,
                )

                heartbeat = master.wait_heartbeat(timeout=5)

                if heartbeat is None:
                    raise TimeoutError("No heartbeat received")

                # Initial request for the telemetry messages used by the HUD.
                request_required_messages(master)

                # Also request the normal ArduPilot data streams. This gives us
                # a fallback if one or more SET_MESSAGE_INTERVAL commands are
                # ignored while the flight controller/radio is still starting.
                master.mav.request_data_stream_send(
                    master.target_system,
                    master.target_component,
                    mavutil.mavlink.MAV_DATA_STREAM_ALL,
                    10,
                    1,
                )

                now = time.monotonic()
                last_message_request_time = now

                with state.lock:
                    state.connected = True
                    state.status = "Telemetry connected"
                    state.last_heartbeat_time = now

                print("Telemetry connected.")

            except Exception as error:
                print(f"Telemetry connection failed: {error}")

                close_master(master)
                master = None

                with state.lock:
                    state.connected = False
                    state.status = "Telemetry disconnected"
                    state.telemetry_link_quality_percent = None
                    state.rc_rssi_percent = None
                    state.last_rc_message_time = 0.0
                    state.rc_failsafe = None

                stop_event.wait(TELEMETRY_RECONNECT_INTERVAL)
                continue

        try:
            # A short timeout prevents this thread from blocking forever.
            msg = master.recv_match(
                blocking=True,
                timeout=0.2,
            )

            now = time.monotonic()

            # Re-send the requested message rates every few seconds.
            # This makes startup order robust:
            #   Python first -> Pixhawk later
            #   Pixhawk first -> Python later
            # both recover automatically without restarting this program.
            if (
                now - last_message_request_time
                >= MESSAGE_REQUEST_RETRY_INTERVAL
            ):
                request_required_messages(master)

                master.mav.request_data_stream_send(
                    master.target_system,
                    master.target_component,
                    mavutil.mavlink.MAV_DATA_STREAM_ALL,
                    10,
                    1,
                )

                last_message_request_time = now

            if msg is not None:
                message_type = msg.get_type()

                telemetry_quality = update_mavlink_packet_quality(
                    msg=msg,
                    now=now,
                    last_seq_by_source=last_seq_by_source,
                    quality_events=quality_events,
                    window_seconds=5.0,
                )

                if telemetry_quality is not None:
                    with state.lock:
                        state.telemetry_link_quality_percent = telemetry_quality

                if message_type == "HEARTBEAT":
                    with state.lock:
                        state.connected = True
                        state.status = "Telemetry connected"
                        state.last_heartbeat_time = now

                elif message_type == "STATUSTEXT":
                    # ArduPilot sends explicit text when Radio Failsafe
                    # becomes active or clears. Use this as an additional
                    # source so the HUD follows the same state that Mission
                    # Planner displays.
                    status_text = str(
                        getattr(msg, "text", "")
                    ).strip()

                    status_text_lower = status_text.lower()

                    if "radio failsafe" in status_text_lower:
                        failsafe_cleared = (
                            "clear" in status_text_lower
                            or "off" in status_text_lower
                            or "recovered" in status_text_lower
                        )

                        failsafe_active = not failsafe_cleared

                        # "PreArm: Radio failsafe on" is explicitly active.
                        if "failsafe on" in status_text_lower:
                            failsafe_active = True

                        with state.lock:
                            state.rc_failsafe = failsafe_active

                            if failsafe_active:
                                state.rc_rssi_percent = None

                        print(
                            "RC failsafe state:",
                            "ACTIVE" if failsafe_active else "CLEARED",
                            f"({status_text})",
                        )

                elif message_type == "SYS_STATUS":
                    voltage = None
                    current = None

                    # voltage_battery is in millivolts.
                    if msg.voltage_battery not in (0, 65535):
                        voltage = msg.voltage_battery / 1000.0

                    # current_battery is in centiamps; -1 means unavailable.
                    if msg.current_battery != -1:
                        current = msg.current_battery / 100.0

                    outputs_enabled = bool(
                        msg.onboard_control_sensors_enabled
                        & mavutil.mavlink.MAV_SYS_STATUS_SENSOR_MOTOR_OUTPUTS
                    )

                    # RC receiver health.
                    #
                    # ArduPilot exposes the RC receiver in SYS_STATUS.
                    # When the receiver is enabled but its health bit drops,
                    # treat that as Radio Failsafe / RC disconnected.
                    rc_receiver_mask = getattr(
                        mavutil.mavlink,
                        "MAV_SYS_STATUS_SENSOR_RC_RECEIVER",
                        65536,  # MAVLink common enum fallback: 1 << 16
                    )

                    rc_receiver_present = bool(
                        msg.onboard_control_sensors_present
                        & rc_receiver_mask
                    )
                    rc_receiver_enabled = bool(
                        msg.onboard_control_sensors_enabled
                        & rc_receiver_mask
                    )
                    rc_receiver_healthy = bool(
                        msg.onboard_control_sensors_health
                        & rc_receiver_mask
                    )

                    with state.lock:
                        if voltage is not None:
                            state.battery_voltage_v = voltage

                        if current is not None:
                            state.total_current_a = current

                        # Only update from SYS_STATUS if ArduPilot says the
                        # RC receiver sensor exists / is enabled.
                        if rc_receiver_present or rc_receiver_enabled:
                            state.rc_failsafe = not rc_receiver_healthy

                            if state.rc_failsafe:
                                # Do not keep displaying the last LQ value
                                # after the transmitter has been lost.
                                state.rc_rssi_percent = None

                    process_safety_state(
                        state,
                        outputs_enabled,
                        now,
                    )

                elif message_type == "BATTERY_STATUS":
                    voltage = None
                    current = None

                    valid_voltages_mv = [
                        value
                        for value in msg.voltages
                        if value not in (0, 65535)
                    ]

                    if valid_voltages_mv:
                        voltage = sum(valid_voltages_mv) / 1000.0

                    if msg.current_battery != -1:
                        current = msg.current_battery / 100.0

                    battery_remaining = getattr(msg, "battery_remaining", -1)
                    if battery_remaining is not None and battery_remaining >= 0:
                        battery_remaining = int(
                            max(0, min(100, battery_remaining))
                        )
                    else:
                        battery_remaining = None

                    with state.lock:
                        if voltage is not None:
                            state.battery_voltage_v = voltage

                        if current is not None:
                            state.total_current_a = current

                        if battery_remaining is not None:
                            state.battery_remaining_percent = battery_remaining

                elif message_type == "VFR_HUD":
                    # groundspeed and climb are metres per second.
                    # Positive climb means ascending; negative means descending.
                    with state.lock:
                        state.ground_speed_m_s = float(msg.groundspeed)
                        state.vertical_speed_m_s = float(msg.climb)

                        heading = float(msg.heading)
                        if heading >= 0:
                            state.heading_deg = heading % 360.0

                elif message_type == "ATTITUDE":
                    # MAVLink ATTITUDE roll/pitch/yaw are in radians.
                    pitch_deg = math.degrees(float(msg.pitch))
                    roll_deg = math.degrees(float(msg.roll))
                    yaw_deg = math.degrees(float(msg.yaw)) % 360.0

                    with state.lock:
                        state.pitch_deg = pitch_deg
                        state.roll_deg = roll_deg
                        state.yaw_deg = yaw_deg

                elif message_type == "GPS_RAW_INT":
                    fix_type = int(msg.fix_type)

                    satellites_visible = getattr(
                        msg,
                        "satellites_visible",
                        None,
                    )

                    if satellites_visible in (None, 255):
                        satellites_visible = None
                    else:
                        satellites_visible = int(satellites_visible)

                    eph = getattr(msg, "eph", None)

                    if eph in (None, 65535):
                        hdop = None
                    else:
                        hdop = float(eph) / 100.0

                    with state.lock:
                        state.gps_fix_type = fix_type
                        state.gps_satellites_visible = satellites_visible
                        state.gps_hdop = hdop

                elif message_type == "GLOBAL_POSITION_INT":
                    # relative_alt is millimetres above the EKF home origin.
                    relative_alt_mm = getattr(msg, "relative_alt", None)

                    if relative_alt_mm is not None:
                        altitude_m = relative_alt_mm / 1000.0

                        # Altitude deadband:
                        # treat values from -0.2 m to +0.2 m as exactly 0 m.
                        if -0.2 <= altitude_m <= 0.2:
                            altitude_m = 0.0

                        with state.lock:
                            state.altitude_m = altitude_m

                elif message_type == "RC_CHANNELS":
                    rc_rssi = getattr(msg, "rssi", None)

                    if rc_rssi in (None, 255):
                        rc_percent = None
                    else:
                        rc_percent = int(round(
                            max(
                                0.0,
                                min(
                                    100.0,
                                    float(rc_rssi) / 254.0 * 100.0,
                                ),
                            )
                        ))

                    with state.lock:
                        state.last_rc_message_time = now

                        # RC_CHANNELS can continue to be transmitted by
                        # ArduPilot even after the physical RC link is lost,
                        # and its RSSI/LQ can retain the last valid value.
                        # Therefore never use it to clear a Radio Failsafe.
                        if (
                            state.rc_failsafe is True
                            or rc_percent is None
                            or rc_percent <= 0
                        ):
                            state.rc_rssi_percent = None
                        else:
                            state.rc_rssi_percent = rc_percent

                elif message_type == "SERVO_OUTPUT_RAW":
                    percentages = []

                    for channel in MOTOR_OUTPUT_CHANNELS:
                        field_name = f"servo{channel}_raw"
                        output_value = getattr(msg, field_name, None)

                        percentages.append(
                            pwm_equivalent_to_percent(output_value)
                        )

                    with state.lock:
                        state.motor_percentages = percentages

            append_log_sample_if_needed(state, now)

            with state.lock:
                heartbeat_age = now - state.last_heartbeat_time

            if heartbeat_age > HEARTBEAT_TIMEOUT:
                print("Heartbeat lost. Telemetry disconnected.")

                with state.lock:
                    state.connected = False
                    state.status = "Telemetry disconnected"
                    state.telemetry_link_quality_percent = None
                    state.rc_rssi_percent = None
                    state.last_rc_message_time = 0.0
                    state.rc_failsafe = None

                close_master(master)
                master = None
                last_message_request_time = 0.0

        except Exception as error:
            print(f"Telemetry error: {error}")

            with state.lock:
                state.connected = False
                state.status = "Telemetry disconnected"
                state.rc_rssi_percent = None
                state.last_rc_message_time = 0.0
                state.rc_failsafe = None

            close_master(master)
            master = None
            last_message_request_time = 0.0

    # Save an unfinished recording when the program closes.
    with state.lock:
        unfinished_recording = state.recording

    if unfinished_recording:
        stop_recording(state)

    close_master(master)


def format_value(value, decimals=1):
    if value is None:
        return "---"

    return f"{value:.{decimals}f}"



def draw_vertical_tape(
    frame,
    value,
    x,
    center_y,
    height,
    major_step,
    minor_step,
    pixels_per_unit,
    label,
    unit,
    ticks_point_right=True,
    fill_negative_region=False,
):  # draw moving HUD line
    """
    Draw a moving vertical HUD tape.

    value:
        Current measured value. If None, zero is used for the scale and the
        value window displays ---.

    x:
        X coordinate of the main vertical scale line.

    ticks_point_right:
        True for the left-side speed tape.
        False for the right-side altitude tape.
    """
    hud_color = (80, 255, 80)
    shadow_color = (0, 0, 0)
    font = cv2.FONT_HERSHEY_SIMPLEX

    top = center_y - height // 2
    bottom = center_y + height // 2
    current_value = 0.0 if value is None else float(value)

    # Main scale line.
    cv2.line(frame, (x, top), (x, bottom), hud_color, 1, cv2.LINE_AA)


    cap_length = max(18, int(frame.shape[1] * 0.05),)                # Horizontal end-cap length.

                                                                        # Make one end of each horizontal line touch the vertical tape.
                                                                        # Left speed tape: caps extend to the left.
                                                                        # Right altitude tape: caps extend to the right.
    if ticks_point_right:
        # Left speed tape: caps extend outward to the left.
        top_cap_start = (x - cap_length, top)
        top_cap_end = (x, top)

        bottom_cap_start = (x - cap_length, bottom)
        bottom_cap_end = (x, bottom)
    else:
        # Right altitude tape: caps extend outward to the right.
        top_cap_start = (x, top)
        top_cap_end = (x + cap_length, top)

        bottom_cap_start = (x, bottom)
        bottom_cap_end = (x + cap_length, bottom)

    cv2.line(
        frame,
        top_cap_start,
        top_cap_end,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    cv2.line(
        frame,
        bottom_cap_start,
        bottom_cap_end,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # Determine which minor tick values are visible.
    value_at_top = current_value + (center_y - top) / pixels_per_unit
    value_at_bottom = current_value - (bottom - center_y) / pixels_per_unit

    first_tick = math.floor(value_at_bottom / minor_step) * minor_step
    last_tick = math.ceil(value_at_top / minor_step) * minor_step

    # For the altitude tape, replace the visible region below 0 m
    # with one continuous solid bar instead of separate tick marks.
    if fill_negative_region:
        zero_y = int(
            round(
                center_y
                - (0.0 - current_value) * pixels_per_unit
            )
        )

        negative_top = max(top, zero_y)
        negative_bottom = bottom

        if negative_top < negative_bottom:                            
            solid_bar_width = 8  # Same width as a major tick.

            if ticks_point_right:
                bar_left = x - solid_bar_width
                bar_right = x
            else:
                bar_left = x
                bar_right = x + solid_bar_width

            cv2.rectangle(
                frame,
                (bar_left, negative_top),
                (bar_right, negative_bottom),
                hud_color,
                -1,
            )

    tick_value = first_tick
    while tick_value <= last_tick + 1e-9:
        tick_y = int(round(center_y - (tick_value - current_value) * pixels_per_unit))

        if top <= tick_y <= bottom:
            if fill_negative_region and tick_value < 0:
                tick_value += minor_step
                continue

            major_ratio = tick_value / major_step
            is_major = abs(major_ratio - round(major_ratio)) < 1e-6
            tick_length = 18 if is_major else 9

            if ticks_point_right:
                                                                            # Left speed tape: ticks and numbers extend outward to the left.
                tick_start = (x - tick_length, tick_y)
                tick_end = (x, tick_y)
                text_x = x - tick_length - 15
            else:
                                                                            # Right altitude tape: ticks and numbers extend outward to the right.
                tick_start = (x, tick_y)
                tick_end = (x + tick_length, tick_y)
                text_x = x + tick_length + 5

            cv2.line(
                frame,
                tick_start,
                tick_end,
                hud_color,
                1,
                cv2.LINE_AA,
            )

            if is_major:

                box_top = center_y - 15
                box_bottom = center_y + 15

                if not (box_top <= tick_y <= box_bottom):

                    tick_label = f"{tick_value:.0f}"

                    cv2.putText(
                        frame,
                        tick_label,
                        (text_x, tick_y + 5),
                        font,
                        0.4,
                        hud_color,
                        1,
                        cv2.LINE_AA,
                    )

        tick_value += minor_step

    # Current-value pointer and box.
    if ticks_point_right:
        # Left speed tape: pointer and value box extend outward to the left.
        pointer = [
            (x + 2, center_y),
            (x - 11, center_y - 8),
            (x - 11, center_y + 8),
        ]
        box_left = x - 55
        box_right = x - 11
    else:
        # Right altitude tape: pointer and value box extend outward to the right.
        pointer = [
            (x - 2, center_y),
            (x + 11, center_y - 8),
            (x + 11, center_y + 8),
        ]
        box_left = x + 11
        box_right = x + 55

    cv2.polylines(
        frame,
        [np.array(pointer, dtype=np.int32)],
        True,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # Draw a semi-transparent black background inside the value box.
    overlay = frame.copy()

    cv2.rectangle(
        overlay,
        (box_left, center_y - 15),
        (box_right, center_y + 15),
        shadow_color,
        -1,
    )

    box_alpha = 0.30  # 0.0 = fully transparent, 1.0 = fully black for altitude tape and speed tape

    cv2.addWeighted(
        overlay,
        box_alpha,
        frame,
        1.0 - box_alpha,
        0,
        frame,
    )

    # Draw the green outline after blending so it remains fully visible.
    cv2.rectangle(
        frame,
        (box_left, center_y - 15),
        (box_right, center_y + 15),
        hud_color,
        1,
        cv2.LINE_AA,
    )

    value_text = "---" if value is None else f"{value:.1f}"
    text_size, _ = cv2.getTextSize(value_text, font, 0.62, 1)
    text_x = box_left + (box_right - box_left - text_size[0]) // 2

    cv2.putText(
        frame,
        value_text,
        (text_x, center_y + 7),
        font,
        0.62,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # Tape title.
    title = f"{label} {unit}"
    title_size, _ = cv2.getTextSize(title, font, 0.45, 1)

    if ticks_point_right:
        # Left title sits outside/left of the speed tape.
        title_x = x - title_size[0] + 3
    else:
        # Right title sits outside/right of the altitude tape.
        title_x = x - 3

    cv2.putText(
        frame,
        title,
        (title_x, top - 10),
        font,
        0.45,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # Show the live value directly below the SPD/ALT title.
    # When no telemetry value is available, display "--".
    title_value_text = "--" if value is None else f"{value:.1f}"

    title_value_size, _ = cv2.getTextSize(
        title_value_text,
        font,
        0.42,
        1,
    )

    # Center the value beneath the title.
    title_value_x = (
        title_x
        + (title_size[0] - title_value_size[0]) // 2
    )
    title_value_y = top + 250

    cv2.putText(
        frame,
        title_value_text,
        (title_value_x, title_value_y),
        font,
        0.42,
        hud_color,
        1,
        cv2.LINE_AA,
    )



def draw_vertical_speed_readout(frame, altitude_tape_x, tape_bottom, vertical_speed_m_s):
    """
    Draw vertical speed below the altitude tape.

    No signal:
        VS: --

    Ascending:
        VS: 1.2 plus a graphical upward arrow

    Descending:
        VS: -0.8 plus a graphical downward arrow
    """
    hud_color = (80, 255, 80)
    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.42
    thickness = 1

    # Place the readout below the ALT tape, aligned with its vertical line.
    text_x = altitude_tape_x
    text_y = min(frame.shape[0] - 12, tape_bottom + 28)

    if vertical_speed_m_s is None:
        text = "VS: --"
        cv2.putText(
            frame,
            text,
            (text_x, text_y),
            font,
            font_scale,
            hud_color,
            thickness,
            cv2.LINE_AA,
        )
        return

    vertical_speed = float(vertical_speed_m_s)
    text = f"VS: {vertical_speed:.1f}"

    cv2.putText(
        frame,
        text,
        (text_x, text_y),
        font,
        font_scale,
        hud_color,
        thickness,
        cv2.LINE_AA,
    )

    # OpenCV's Hershey font does not reliably support Unicode arrows,
    # so draw the arrow as lines.
    text_size, _ = cv2.getTextSize(
        text,
        font,
        font_scale,
        thickness,
    )

    arrow_x = text_x + text_size[0] + 10
    arrow_center_y = text_y - 5
    arrow_length = 14
    arrow_head = 5

    # Small deadband: near-zero climb shows the number without an arrow.
    if vertical_speed > 0.05:
        arrow_top = arrow_center_y - arrow_length // 2
        arrow_bottom = arrow_center_y + arrow_length // 2

        cv2.line(
            frame,
            (arrow_x, arrow_bottom),
            (arrow_x, arrow_top),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )
        cv2.line(
            frame,
            (arrow_x, arrow_top),
            (arrow_x - arrow_head, arrow_top + arrow_head),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )
        cv2.line(
            frame,
            (arrow_x, arrow_top),
            (arrow_x + arrow_head, arrow_top + arrow_head),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )

    elif vertical_speed < -0.05:
        arrow_top = arrow_center_y - arrow_length // 2
        arrow_bottom = arrow_center_y + arrow_length // 2

        cv2.line(
            frame,
            (arrow_x, arrow_top),
            (arrow_x, arrow_bottom),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )
        cv2.line(
            frame,
            (arrow_x, arrow_bottom),
            (arrow_x - arrow_head, arrow_bottom - arrow_head),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )
        cv2.line(
            frame,
            (arrow_x, arrow_bottom),
            (arrow_x + arrow_head, arrow_bottom - arrow_head),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )



def draw_primary_hud_tapes(frame, speed_m_s, altitude_m, vertical_speed_m_s):
    """
    Draw the two main HUD tapes requested by the user:
    left = horizontal/ground speed, right = relative altitude.
    """
    frame_height, frame_width = frame.shape[:2]

    center_y = int(frame_height*0.4)
    tape_height = max(180, int(frame_height * 0.5))               # airspeed tap height and altitude tap height

    
    left_x = int(frame_width * 0.28)                               # left side 20% from the edge
    right_x = int(frame_width * 0.72)                              # right side 80% from the edge

    draw_vertical_tape(                                            # draw speed tape
        frame=frame,
        value=speed_m_s,
        x=left_x,
        center_y=center_y,
        height=tape_height,
        major_step=5.0,
        minor_step=1.0,
        pixels_per_unit=14.0,
        label="SPD",
        unit="m/s",
        ticks_point_right=True,
    )

    draw_vertical_tape(                                            # draw altitude tape
        frame=frame,
        value=altitude_m,
        x=right_x,
        center_y=center_y,
        height=tape_height,
        major_step=1,
        minor_step=0.2,
        pixels_per_unit=70,
        label="ALT",
        unit="m",
        ticks_point_right=False,
        fill_negative_region=True,
    )

    altitude_tape_bottom = center_y + tape_height // 2

    draw_vertical_speed_readout(
        frame=frame,
        altitude_tape_x=right_x,
        tape_bottom=altitude_tape_bottom,
        vertical_speed_m_s=vertical_speed_m_s,
    )




def draw_pitch_ladder(frame, pitch_deg, roll_deg):
    """
    Draw pitch ladder with opposite roll compensation.

    If aircraft rolls right +10 deg, the pitch ladder rotates left -10 deg.
    Rotation center is the fixed aircraft reference symbol.
    """
    frame_height, frame_width = frame.shape[:2]

    hud_color = (80, 255, 80)
    font = cv2.FONT_HERSHEY_SIMPLEX

    center_x = frame_width // 2
    reference_y = int(frame_height * 0.39)

    current_pitch = 0.0 if pitch_deg is None else float(pitch_deg)
    current_roll = 0.0 if roll_deg is None else float(roll_deg)

    # Opposite direction to aircraft roll.
    ladder_roll_rad = math.radians(-current_roll)
    cos_r = math.cos(ladder_roll_rad)
    sin_r = math.sin(ladder_roll_rad)

    def rotate_point(x, y):
        dx = x - center_x
        dy = y - reference_y

        rx = center_x + dx * cos_r - dy * sin_r
        ry = reference_y + dx * sin_r + dy * cos_r

        return int(round(rx)), int(round(ry))

    pixels_per_degree = max(3.5, frame_height * 0.018)
    pitch_step = 5.0

    tape_center_y = int(frame_height * 0.4)
    tape_height = max(180, int(frame_height * 0.5))
    tape_top = tape_center_y - tape_height // 2

    distance_to_top = reference_y - tape_top
    visible_top = max(0, tape_top)
    visible_bottom = min(
        frame_height - 1,
        reference_y + distance_to_top,
    )

    normal_half_width = max(55, int(frame_width * 0.095))
    horizon_half_width = max(85, int(frame_width * 0.2))
    center_gap = max(24, int(frame_width * 0.035))

    left_limit = int(frame_width * 0.34)
    right_limit = int(frame_width * 0.66)

    pitch_at_top = (
        current_pitch
        + (reference_y - visible_top) / pixels_per_degree
    )

    pitch_at_bottom = (
        current_pitch
        + (reference_y - visible_bottom) / pixels_per_degree
    )

    lowest_visible_pitch = min(pitch_at_top, pitch_at_bottom)
    highest_visible_pitch = max(pitch_at_top, pitch_at_bottom)

    first_mark = (
        math.ceil(lowest_visible_pitch / pitch_step)
        * pitch_step
    )

    last_mark = (
        math.floor(highest_visible_pitch / pitch_step)
        * pitch_step
    )

    mark_deg = first_mark

    while mark_deg <= last_mark + 1e-9:
        base_y = int(round(
            reference_y
            + (current_pitch - mark_deg) * pixels_per_degree
        ))

        if visible_top <= base_y <= visible_bottom:
            if abs(mark_deg) < 1e-9:
                half_width = horizon_half_width
            else:
                half_width = normal_half_width

            left_x = max(left_limit, center_x - half_width)
            right_x = min(right_limit, center_x + half_width)

            left_outer = rotate_point(left_x, base_y)
            left_inner = rotate_point(center_x - center_gap, base_y)
            right_inner = rotate_point(center_x + center_gap, base_y)
            right_outer = rotate_point(right_x, base_y)

            cv2.line(
                frame,
                left_outer,
                left_inner,
                hud_color,
                1,
                cv2.LINE_AA,
            )

            cv2.line(
                frame,
                right_inner,
                right_outer,
                hud_color,
                1,
                cv2.LINE_AA,
            )

            hook_height = max(4, int(frame_height * 0.010))

            if mark_deg > 0:
                left_hook_end = rotate_point(
                    left_x,
                    min(base_y + hook_height, visible_bottom),
                )
                right_hook_end = rotate_point(
                    right_x,
                    min(base_y + hook_height, visible_bottom),
                )

                cv2.line(
                    frame,
                    left_outer,
                    left_hook_end,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

                cv2.line(
                    frame,
                    right_outer,
                    right_hook_end,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

            elif mark_deg < 0:
                left_hook_end = rotate_point(
                    left_x,
                    max(base_y - hook_height, visible_top),
                )
                right_hook_end = rotate_point(
                    right_x,
                    max(base_y - hook_height, visible_top),
                )

                cv2.line(
                    frame,
                    left_outer,
                    left_hook_end,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

                cv2.line(
                    frame,
                    right_outer,
                    right_hook_end,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

            # Labels move with the ladder, but stay upright.
            if abs(mark_deg) >= 1e-9:
                label = f"{int(round(mark_deg)):+d}"
                text_size, _ = cv2.getTextSize(
                    label,
                    font,
                    0.38,
                    1,
                )

                left_label = rotate_point(left_x - 8, base_y)
                right_label = rotate_point(right_x + 8, base_y)

                cv2.putText(
                    frame,
                    label,
                    (left_label[0] - text_size[0], left_label[1] + 5),
                    font,
                    0.38,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

                cv2.putText(
                    frame,
                    label,
                    (right_label[0], right_label[1] + 5),
                    font,
                    0.38,
                    hud_color,
                    1,
                    cv2.LINE_AA,
                )

        mark_deg += pitch_step




def draw_lower_center_symbol(frame):
    """
    Draw a sharp green reference symbol in the lower half of the screen.

    The symbol stays horizontally centered and scales with the frame size.
    A thin dark outline plus a bright center line keeps the symbol clear
    without the blurry glow produced by a thick anti-aliased line.
    """
    frame_height, frame_width = frame.shape[:2]

    outline_color = (0, 110, 0)
    hud_color = (80, 255, 80)

    
    center_x = frame_width // 2
    center_y = int(frame_height * 0.39)                                      # aircraft position

    # Scale the symbol with the video resolution.
    symbol_width = max(80, int(frame_width * 0.16))
    notch_width = max(24, int(symbol_width * 0.28))
    notch_depth = max(7, int(frame_height * 0.022))

    # Short sloped transitions make the corners smoother and clearer.
    slope_width = max(8, int(symbol_width * 0.06))

    left_x = center_x - symbol_width // 2
    right_x = center_x + symbol_width // 2

    notch_left_x = center_x - notch_width // 2
    notch_right_x = center_x + notch_width // 2

    points = np.array(
        [
            (left_x, center_y),
            (notch_left_x - slope_width, center_y),
            (notch_left_x, center_y + notch_depth),
            (notch_right_x, center_y + notch_depth),
            (notch_right_x + slope_width, center_y),
            (right_x, center_y),
        ],
        dtype=np.int32,
    )

    # Dark outline: LINE_8 prevents a wide blurry anti-aliased glow.
    cv2.polylines(
        frame,
        [points],
        False,
        outline_color,
        3,
        cv2.LINE_8,
    )

    # Bright one-pixel center line.
    cv2.polylines(
        frame,
        [points],
        False,
        hud_color,
        1,
        cv2.LINE_AA,
    )


def draw_rotating_compass(
    frame,
    heading_deg,
    speed_m_s,
    gps_fix_type,
    gps_satellites_visible,
    gps_hdop,
):
    """
    Draw a rotating 360-degree compass rose at the bottom of the HUD.

    Design:
        - The aircraft itself is fixed relative to the screen.
        - A fixed inverted triangle at the top of the compass represents
          the aircraft's current heading reference.
        - The compass rose rotates with respect to the real world.
        - North = 000 deg, East = 090 deg, South = 180 deg, West = 270 deg.
        - One tick is drawn for every 1 degree.
        - Longer ticks are used every 5, 10, 30 and 90 degrees.

    Rotation logic:
        If aircraft heading = 0 deg:
            N is directly beneath the fixed triangle.

        If aircraft heading = 90 deg:
            E is directly beneath the fixed triangle.

        Therefore the compass rose is rotated by:
            world_angle - aircraft_heading
    """
    frame_height, frame_width = frame.shape[:2]

    hud_color = (80, 255, 80)
    shadow_color = (0, 0, 0)
    font = cv2.FONT_HERSHEY_SIMPLEX

    # Use zero only as a visual fallback before telemetry becomes available.
    heading = 0.0 if heading_deg is None else float(heading_deg) % 360.0

    # ---------------------------------------------------------
    # Compass geometry
    # ---------------------------------------------------------
    center_x = frame_width // 2

    # Compass size.
    radius = max(55, int(min(frame_width, frame_height) * 0.15))

    # Compass vertical position.
    #
    # The compass is intentionally allowed to extend below the image.
    # COMPASS_DOWN_OFFSET controls how far it moves downward.
    #
    # Example:
    #   0  -> bottom of compass just touches bottom of frame
    #   40 -> bottom 40 px of compass is outside the frame
    center_y = (
        frame_height
        - radius
        + COMPASS_DOWN_OFFSET
    )


    # ---------------------------------------------------------
    # Draw all 360 one-degree ticks.
    #
    # Screen-angle convention:
    #     0 deg is straight up on the screen.
    #     positive angle rotates clockwise.
    #
    # For a world direction d:
    #     screen_angle = d - heading
    #
    # Example:
    #     heading = 90
    #     East = 90
    #     screen_angle = 0
    # Therefore E appears at the top under the aircraft pointer.
    # ---------------------------------------------------------
    for world_deg in range(0,360,10):
        screen_deg = (world_deg - heading) % 360.0
        angle_rad = math.radians(screen_deg)

        # Tick length hierarchy.
        if world_deg % 90 == 0:
            tick_length = max(12, int(radius * 0.16))
            thickness = 2
        elif world_deg % 30 == 0:
            tick_length = max(10, int(radius * 0.13))
            thickness = 1
        elif world_deg % 10 == 0:
            tick_length = max(8, int(radius * 0.10))
            thickness = 1
        elif world_deg % 5 == 0:
            tick_length = max(6, int(radius * 0.075))
            thickness = 1
        else:
            tick_length = max(3, int(radius * 0.045))
            thickness = 1

        outer_radius = radius
        inner_radius = radius - tick_length

        outer_x = int(round(
            center_x + outer_radius * math.sin(angle_rad)
        ))
        outer_y = int(round(
            center_y - outer_radius * math.cos(angle_rad)
        ))

        inner_x = int(round(
            center_x + inner_radius * math.sin(angle_rad)
        ))
        inner_y = int(round(
            center_y - inner_radius * math.cos(angle_rad)
        ))

        cv2.line(
            frame,
            (inner_x, inner_y),
            (outer_x, outer_y),
            hud_color,
            thickness,
            cv2.LINE_AA,
        )

    # ---------------------------------------------------------
    # Cardinal labels: N, E, S, W.
    # These labels rotate with the real-world compass rose.
    # ---------------------------------------------------------
    cardinal_directions = {
        0: "N",
        90: "E",
        180: "S",
        270: "W",
    }

    label_radius = radius - max(24, int(radius * 0.25))

    for world_deg, label in cardinal_directions.items():
        screen_deg = (world_deg - heading) % 360.0
        angle_rad = math.radians(screen_deg)

        label_center_x = int(round(
            center_x + label_radius * math.sin(angle_rad)
        ))
        label_center_y = int(round(
            center_y - label_radius * math.cos(angle_rad)
        ))

        text_size, _ = cv2.getTextSize(
            label,
            font,
            0.52,
            1,
        )

        cv2.putText(
            frame,
            label,
            (
                label_center_x - text_size[0] // 2,
                label_center_y + text_size[1] // 2,
            ),
            font,
            0.52,
            hud_color,
            1,
            cv2.LINE_AA,
        )

    # ---------------------------------------------------------
    # Numerical labels every 30 degrees.
    #
    # Cardinal points keep N/E/S/W instead of numbers.
    # Other labels use aviation-style tens:
    #     030 -> 3
    #     060 -> 6
    #     120 -> 12
    # etc.
    # ---------------------------------------------------------
    number_radius = int(radius * 0.75)

    for world_deg in range(0, 360, 30):
        if world_deg in cardinal_directions:
            continue

        screen_deg = (world_deg - heading) % 360.0
        angle_rad = math.radians(screen_deg)

        number_center_x = int(round(
            center_x + number_radius * math.sin(angle_rad)
        ))
        number_center_y = int(round(
            center_y - number_radius * math.cos(angle_rad)
        ))

        number_text = str(world_deg // 10)

        text_size, _ = cv2.getTextSize(
            number_text,
            font,
            0.34,
            1,
        )

        cv2.putText(
            frame,
            number_text,
            (
                number_center_x - text_size[0] // 2,
                number_center_y + text_size[1] // 2,
            ),
            font,
            0.34,
            hud_color,
            1,
            cv2.LINE_AA,
        )

    # ---------------------------------------------------------
    # Fixed inverted triangle at the top center.
    #
    # This DOES NOT rotate. It represents the aircraft heading line.
    # The compass rose moves underneath it.
    # ---------------------------------------------------------
    triangle_top_y = center_y - radius - 10
    triangle_half_width = max(6, int(radius * 0.07))
    triangle_height = max(8, int(radius * 0.09))

    pointer_points = np.array(
        [
            (center_x - triangle_half_width, triangle_top_y),
            (center_x + triangle_half_width, triangle_top_y),
            (center_x, triangle_top_y + triangle_height),
        ],
        dtype=np.int32,
    )

    cv2.polylines(
        frame,
        [pointer_points],
        True,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # ---------------------------------------------------------
    # Left-side SPD and HDG readouts.
    # ---------------------------------------------------------
    if heading_deg is None:
        heading_text = "HDG ---"
    else:
        heading_text = f"HDG {int(round(heading)) % 360:03d}"

    if speed_m_s is None:
        speed_text = "SPD ---"
    else:
        speed_text = f"SPD {float(speed_m_s):.1f}"

    heading_size, _ = cv2.getTextSize(
        heading_text,
        font,
        0.42,
        1,
    )

    speed_size, _ = cv2.getTextSize(
        speed_text,
        font,
        0.42,
        1,
    )

    max_text_width = max(
        heading_size[0],
        speed_size[0],
    )

    # Put both labels on the LEFT side of the compass.
    text_x = center_x - radius - max_text_width - max(
        15,
        int(frame_width * 0.02),
    )
    text_x = max(5, text_x)

    # HDG remains vertically centered beside the compass.
    text_up_offset = 30
    heading_text_y = center_y + heading_size[1] // 2 - text_up_offset

    # SPD is directly above HDG.
    line_spacing = max(
        22,
        int(frame_height * 0.045),
    )
    speed_text_y = heading_text_y - line_spacing

    cv2.putText(
        frame,
        speed_text,
        (
            text_x,
            speed_text_y,
        ),
        font,
        0.42,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    cv2.putText(
        frame,
        heading_text,
        (
            text_x,
            heading_text_y,
        ),
        font,
        0.42,
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # ---------------------------------------------------------
    # Right-side GPS readouts, symmetrical to SPD / HDG.
    #
    # GPS is binary GREEN / RED:
    # GREEN only when:
    #   fix_type >= 3
    #   satellites > 3
    #   HDOP <= 2.0
    # Otherwise RED.
    # ---------------------------------------------------------
    fix_text = gps_fix_type_to_text(gps_fix_type)

    if gps_fix_type is None:
        gps_text = "GPS: NO GPS"
    else:
        gps_text = f"GPS: {fix_text}"

    satellites_text = (
        "--"
        if gps_satellites_visible is None
        else str(gps_satellites_visible)
    )

    hdop_text = (
        "--"
        if gps_hdop is None
        else f"{gps_hdop:.1f}"
    )

    sat_text = f"SAT: {satellites_text}  HDOP {hdop_text}"

    gps_good = (
        gps_fix_type is not None
        and gps_fix_type >= GPS_MIN_FIX_TYPE
        and gps_satellites_visible is not None
        and gps_satellites_visible > GPS_MIN_SATELLITES_EXCLUSIVE
        and gps_hdop is not None
        and gps_hdop <= GPS_MAX_HDOP
    )

    gps_color = (
        (80, 255, 80) if gps_good
        else (0, 0, 255)
    )

    gps_size, _ = cv2.getTextSize(
        gps_text,
        font,
        0.42,
        1,
    )

    sat_size, _ = cv2.getTextSize(
        sat_text,
        font,
        0.42,
        1,
    )

    right_text_width = max(
        gps_size[0],
        sat_size[0],
    )

    right_text_x = center_x + radius + max(
        15,
        int(frame_width * 0.02),
    )

    right_text_x = min(
        right_text_x,
        frame_width - right_text_width - 5,
    )

    cv2.putText(
        frame,
        gps_text,
        (right_text_x, speed_text_y),
        font,
        0.42,
        gps_color,
        1,
        cv2.LINE_AA,
    )

    cv2.putText(
        frame,
        sat_text,
        (right_text_x, heading_text_y),
        font,
        0.42,
        gps_color,
        1,
        cv2.LINE_AA,
    )

     # ---------------------------------------------------------
    # Copter symbol at compass center
    # Small circle + X
    # ---------------------------------------------------------
    copter_radius = max(5, int(radius * 0.045))

    # Small center circle
    diamond_size = max(6, int(radius * 0.03))

    diamond_points = np.array(
        [
            (center_x, center_y - diamond_size),      # top
            (center_x + diamond_size, center_y),      # right
            (center_x, center_y + diamond_size),      # bottom
            (center_x - diamond_size, center_y),      # left
        ],
        dtype=np.int32,
    )

    cv2.polylines(
        frame,
        [diamond_points],
        True,
        hud_color,
        1,
        cv2.LINE_AA,
    )
    # X inside the circle
    x_size = max(12, int(copter_radius * 1))

    cv2.line(
        frame,
        (center_x - x_size, center_y - x_size),
        (center_x + x_size, center_y + x_size),
        hud_color,
        1,
        cv2.LINE_AA,
    )

    cv2.line(
        frame,
        (center_x + x_size, center_y - x_size),
        (center_x - x_size, center_y + x_size),
        hud_color,
        1,
        cv2.LINE_AA,
    )

    # ---------------------------------------------------------
    # GPS warning at the compass center.
    #
    # GPS_RAW_INT fix_type:
    #   0 = no GPS
    #   1 = no fix
    #   2 = 2D fix
    #   3+ = valid 3D (or better) fix
    #
    # Missing GPS data and anything below 3D fix are treated as NOT FIX.
    #
    # The warning is split into TWO lines so the existing center copter
    # symbol (diamond + X) remains visible and is NOT covered.
    # ---------------------------------------------------------
    if gps_fix_type is None or gps_fix_type < 3:
        warning_color = (0, 0, 255)
        warning_font_scale = 0.48
        warning_thickness = 1

        top_text = " GPS NOT FIX"
        

        top_size, _ = cv2.getTextSize(
            top_text,
            font,
            warning_font_scale,
            warning_thickness,
        )



        # Leave a clear gap around the existing copter symbol.
        # Increase this value if you want the red text farther away
        # from the diamond/X.
        symbol_clearance = max(
            18,
            x_size + diamond_size + 6,
        )

        # "GPS" above the center symbol.
        top_x = center_x - top_size[0] // 2
        top_y = center_y - symbol_clearance

        cv2.putText(
            frame,
            top_text,
            (top_x, top_y),
            font,
            warning_font_scale,
            warning_color,
            warning_thickness,
            cv2.LINE_AA,
        )



def gps_fix_type_to_text(fix_type):
    if fix_type is None:
        return "--"

    fix_names = {
        0: "NO GPS",
        1: "NO FIX",
        2: "2D",
        3: "3D",
        4: "DGPS",
        5: "RTK FLOAT",
        6: "RTK FIX",
    }

    return fix_names.get(
        int(fix_type),
        str(int(fix_type)),
    )


def draw_lower_left_status_panel(
    frame,
    connected,
    telemetry_link_quality_percent,
    rc_rssi_percent,
    rc_failsafe,
    total_current_a,
    battery_voltage_v,
    battery_remaining_percent,
    test_alert_mode=0,
):
    """
    Lower-left panel with three-level colors:

    TEL / RC:
        GREEN  >= 40%
        YELLOW 20% .. <40%
        RED    <20%

    CUR:
        GREEN  <=60 A
        YELLOW >60 A .. <=90 A
        RED    >90 A

    BAT:
        GREEN  >=14.0 V
        YELLOW 13.2 V .. <14.0 V
        RED    <13.2 V
    """
    frame_height, frame_width = frame.shape[:2]

    green = (80, 255, 80)
    yellow = (0, 255, 255)
    red = (0, 0, 255)

    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.42
    thickness = 1

    text_x = (
        HUD_SAFE_MARGIN
        + max(10, int(frame_width * 0.015))
        + 40
    )

    line_spacing = max(
        22,
        int(frame_height * 0.045),
    )

    fourth_line_y = (
        frame_height
        - HUD_SAFE_MARGIN
        - 12
        - 90
    )

    first_line_y = fourth_line_y - 3 * line_spacing

    # TEL
    if connected and telemetry_link_quality_percent is not None:
        tel_text = f"TEL: {telemetry_link_quality_percent}%"

        if telemetry_link_quality_percent < TEL_CRITICAL_PERCENT:
            tel_color = red
        elif telemetry_link_quality_percent < TEL_WARNING_PERCENT:
            tel_color = yellow
        else:
            tel_color = green
    else:
        tel_text = "TEL: --"
        tel_color = green

    # RC
    # When ArduPilot reports Radio Failsafe, or the reported receiver
    # signal value has dropped to 0, treat the RC link as disconnected.
    if (
        connected
        and rc_failsafe is not True
        and rc_rssi_percent is not None
        and rc_rssi_percent > 0
    ):
        rc_text = f"RC : {rc_rssi_percent}%"

        if rc_rssi_percent < RC_CRITICAL_PERCENT:
            rc_color = red
        elif rc_rssi_percent < RC_WARNING_PERCENT:
            rc_color = yellow
        else:
            rc_color = green
    else:
        rc_text = "RC : --"
        rc_color = green

    # CUR
    if connected and total_current_a is not None:
        current_text = f"CUR: {total_current_a:.1f}A"

        if total_current_a > CURRENT_CRITICAL_A:
            current_color = red
        elif total_current_a > CURRENT_WARNING_A:
            current_color = yellow
        else:
            current_color = green
    else:
        current_text = "CUR: --A"
        current_color = green

    # BAT
    if connected and battery_voltage_v is not None:
        voltage_text = f"{battery_voltage_v:.1f}V"

        if battery_voltage_v < BAT_VOLTAGE_CRITICAL:
            battery_color = red
        elif battery_voltage_v < BAT_VOLTAGE_WARNING:
            battery_color = yellow
        else:
            battery_color = green
    else:
        voltage_text = "--V"
        battery_color = green

    if connected and battery_remaining_percent is not None:
        battery_percent_text = f"{battery_remaining_percent}%"
    else:
        battery_percent_text = "--%"

    battery_text = f"BAT: {voltage_text}  {battery_percent_text}"

    # Manual alert-test override.
    # Keep the real displayed values; only force the alert state/color.
    if test_alert_mode == 1:
        tel_color = yellow
        rc_color = yellow
        current_color = yellow
        battery_color = yellow

    elif test_alert_mode == 2:
        tel_color = red
        rc_color = red
        current_color = red
        battery_color = red

    lines = (
        (tel_text, tel_color),
        (rc_text, rc_color),
        (current_text, current_color),
        (battery_text, battery_color),
    )

    for line_index, (text, color) in enumerate(lines):
        text_y = first_line_y + line_index * line_spacing

        cv2.putText(
            frame,
            text,
            (text_x, text_y),
            font,
            font_scale,
            color,
            thickness,
            cv2.LINE_AA,
        )


def draw_system_warnings(
    frame,
    connected,
    telemetry_link_quality_percent,
    rc_rssi_percent,
    total_current_a,
    battery_voltage_v,
    test_alert_mode=0,
):
    """
    Upper-left warning messages.

    Warning range:
        steady YELLOW

    Critical range:
        flashing RED
    """
    # In normal mode, warnings require a live telemetry connection.
    # In manual test modes 1/2, warnings must still be visible even when
    # telemetry is disconnected.
    if not connected and test_alert_mode == 0:
        return

    yellow = (0, 255, 255)
    red = (0, 0, 255)

    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.52
    thickness = 1

    warnings = []

    # Manual test mode overrides the live threshold logic.
    if test_alert_mode == 1:
        warnings = [
            ("Telemetry is low", "warning"),
            ("RC is low", "warning"),
            ("Current is high", "warning"),
            ("Battery is low", "warning"),
        ]

    elif test_alert_mode == 2:
        warnings = [
            ("Telemetry is low", "critical"),
            ("RC is low", "critical"),
            ("Current is high", "critical"),
            ("Battery is low", "critical"),
        ]

    # Normal live threshold logic.
    elif test_alert_mode == 0 and telemetry_link_quality_percent is not None:
        if telemetry_link_quality_percent < TEL_CRITICAL_PERCENT:
            warnings.append(("Telemetry is low", "critical"))
        elif telemetry_link_quality_percent < TEL_WARNING_PERCENT:
            warnings.append(("Telemetry is low", "warning"))

    # RC
    if test_alert_mode == 0 and rc_rssi_percent is not None:
        if rc_rssi_percent < RC_CRITICAL_PERCENT:
            warnings.append(("RC is low", "critical"))
        elif rc_rssi_percent < RC_WARNING_PERCENT:
            warnings.append(("RC is low", "warning"))

    # CUR
    if test_alert_mode == 0 and total_current_a is not None:
        if total_current_a > CURRENT_CRITICAL_A:
            warnings.append(("Current is high", "critical"))
        elif total_current_a > CURRENT_WARNING_A:
            warnings.append(("Current is high", "warning"))

    # BAT
    if test_alert_mode == 0 and battery_voltage_v is not None:
        if battery_voltage_v < BAT_VOLTAGE_CRITICAL:
            warnings.append(("Battery is low", "critical"))
        elif battery_voltage_v < BAT_VOLTAGE_WARNING:
            warnings.append(("Battery is low", "warning"))

    if not warnings:
        return

    blink_on = (
        time.monotonic() % WARNING_BLINK_PERIOD_S
    ) < (WARNING_BLINK_PERIOD_S / 2.0)

    text_x = HUD_SAFE_MARGIN + 10
    first_y = HUD_SAFE_MARGIN + 20
    line_spacing = max(
        24,
        int(frame.shape[0] * 0.045),
    )

    visible_index = 0

    for text, severity in warnings:
        if severity == "critical":
            if not blink_on:
                continue
            color = red
        else:
            color = yellow

        text_y = first_y + visible_index * line_spacing

        cv2.putText(
            frame,
            text,
            (text_x, text_y),
            font,
            font_scale,
            color,
            thickness,
            cv2.LINE_AA,
        )

        visible_index += 1


def draw_disconnect_messages(
    frame,
    camera_connected,
    telemetry_connected,
    rc_connected,
):
    """
    Draw connection-loss messages over the middle of the pitch ladder.

    Each message is shown only while that connection is missing:
        USB Camera Disconnect
        Telemetry Disconnect
        RC Signal Disconnect
    """
    frame_height, frame_width = frame.shape[:2]

    font = cv2.FONT_HERSHEY_SIMPLEX
    font_scale = 0.5
    thickness = 1
    warning_color = (0, 0, 255)  # red
    shadow_color = (0, 0, 0)

    # Keep the three rows at fixed locations, so each warning has its own
    # position and simply disappears when that connection becomes available.
    messages = [
        ("USB Camera Disconnect", not camera_connected, -1),
        ("Telemetry Disconnect", not telemetry_connected, 0),
        ("RC Signal Disconnect", not rc_connected, 1),
    ]

    center_x = frame_width // 2
    center_y = int(frame_height * 0.39)
    line_spacing = max(30, int(frame_height * 0.1))

    for message, visible, row_offset in messages:
        if not visible:
            continue

        text_size, _ = cv2.getTextSize(
            message,
            font,
            font_scale,
            thickness,
        )

        text_x = center_x - text_size[0] // 2
        text_y = center_y + row_offset * line_spacing


        cv2.putText(
            frame,
            message,
            (text_x, text_y),
            font,
            font_scale,
            warning_color,
            thickness,
            cv2.LINE_AA,
        )

def heading_to_cardinal(heading_deg):
    if heading_deg is None:
        return "---"

    directions = ("N", "NE", "E", "SE", "S", "SW", "W", "NW")
    index = int((heading_deg + 22.5) // 45) % 8
    return directions[index]


def draw_telemetry(frame, state, test_alert_mode=0):
    with state.lock:
        connected = state.connected
        status = state.status
        voltage = state.battery_voltage_v
        battery_remaining = state.battery_remaining_percent
        current = state.total_current_a
        telemetry_link_quality = state.telemetry_link_quality_percent
        rc_rssi = state.rc_rssi_percent
        rc_failsafe = state.rc_failsafe
        altitude = state.altitude_m
        ground_speed = state.ground_speed_m_s
        vertical_speed = state.vertical_speed_m_s
        heading = state.heading_deg
        gps_fix_type = state.gps_fix_type
        gps_satellites_visible = state.gps_satellites_visible
        gps_hdop = state.gps_hdop
        yaw = state.yaw_deg
        pitch = state.pitch_deg
        roll = state.roll_deg
        motor_percentages = list(state.motor_percentages)
        recording = state.recording
        record_count = len(state.records)

    # =========================================================
    # DRAW ALL HUD GRAPHICS ON A TEMPORARY FRAME
    #
    # frame:
    #     original analog video
    #
    # hud_frame:
    #     analog video + ALL green HUD graphics
    #
    # Nothing is drawn directly onto frame until the safe margin
    # has been applied.
    # =========================================================
    hud_frame = frame.copy()

    # Main speed / altitude tapes.
    draw_primary_hud_tapes(
        hud_frame,
        speed_m_s=ground_speed if connected else None,
        altitude_m=altitude if connected else None,
        vertical_speed_m_s=vertical_speed if connected else None,
    )

    # Central pitch ladder.
    draw_pitch_ladder(
        hud_frame,
        pitch_deg=pitch if connected else None,
        roll_deg=roll if connected else None,
    )

    # Aircraft reference / attitude symbol.
    draw_lower_center_symbol(
        hud_frame
    )

    # Rotating compass + SPD + HDG.
    #
    # IMPORTANT:
    # The compass is also drawn on hud_frame so it is clipped by
    # HUD_SAFE_MARGIN just like every other green HUD element.
    draw_rotating_compass(
        hud_frame,
        heading_deg=yaw if connected else None,
        speed_m_s=ground_speed if connected else None,
        gps_fix_type=gps_fix_type if connected else None,
        gps_satellites_visible=(
            gps_satellites_visible if connected else None
        ),
        gps_hdop=gps_hdop if connected else None,
    )

    # Lower-left telemetry / RC / GPS / battery status.
    draw_lower_left_status_panel(
    hud_frame,
    connected=connected,
    telemetry_link_quality_percent=telemetry_link_quality,
    rc_rssi_percent=rc_rssi,
    rc_failsafe=rc_failsafe,
    total_current_a=current,
    battery_voltage_v=voltage,
    battery_remaining_percent=battery_remaining,
    test_alert_mode=test_alert_mode,
    )
    # =========================================================
    # SAFE BORDER / MARGIN
    #
    # Only copy the central region from hud_frame back onto frame.
    #
    # The outer border therefore remains completely untouched
    # original analog video.
    # =========================================================
    frame_height, frame_width = frame.shape[:2]

    margin = int(HUD_SAFE_MARGIN)

    # Keep the margin inside valid slicing limits.
    max_margin_x = max(0, frame_width // 2 - 1)
    max_margin_y = max(0, frame_height // 2 - 1)

    margin = max(
        0,
        min(
            margin,
            max_margin_x,
            max_margin_y,
        ),
    )

    if margin == 0:
        # No protected margin.
        frame[:, :] = hud_frame

    else:
        # Copy only the safe center region.
        frame[
            margin:frame_height - margin,
            margin:frame_width - margin,
        ] = hud_frame[
            margin:frame_height - margin,
            margin:frame_width - margin,
        ]
    

    # Upper-left TEL / RC / CUR / BAT alerts.
    draw_system_warnings(
        frame,
        connected=connected,
        telemetry_link_quality_percent=telemetry_link_quality,
        rc_rssi_percent=rc_rssi,
        total_current_a=current,
        battery_voltage_v=voltage,
        test_alert_mode=test_alert_mode,
    )


def main():
    cv2.namedWindow(
        WINDOW_NAME,
        cv2.WINDOW_NORMAL,
    )

    no_camera_screen = create_no_camera_screen()

    telemetry_state = TelemetryState()
    stop_event = threading.Event()

    telemetry_thread = threading.Thread(
        target=telemetry_worker,
        args=(telemetry_state, stop_event),
        daemon=True,
    )

    telemetry_thread.start()

    cap = None
    last_camera_connection_attempt = 0.0

    # 0 = normal, 1 = force yellow warnings, 2 = force red warnings.
    test_alert_mode = 0

    try:
        while True:
            current_time = time.monotonic()

            if cap is None:
                display_frame = no_camera_screen.copy()
                camera_connected = False

                if (
                    current_time
                    - last_camera_connection_attempt
                    >= CAMERA_RECONNECT_INTERVAL
                ):
                    last_camera_connection_attempt = current_time
                    print("Checking for USB camera...")

                    cap = open_usb_camera()

                    if cap is not None:
                        print("USB camera detected.")
                        camera_connected = True

            else:
                ret, frame = cap.read()

                if ret and frame is not None:
                    display_frame = frame
                    camera_connected = True

                else:
                    print("USB camera disconnected.")

                    cap.release()
                    cap = None
                    camera_connected = False
                    display_frame = no_camera_screen.copy()

            # Draw the normal HUD first.
            draw_telemetry(
                display_frame,
                telemetry_state,
                test_alert_mode=test_alert_mode,
            )

            # Determine the current Telemetry and RC connection states.
            #
            # IMPORTANT:
            # RC_CHANNELS may keep arriving after the transmitter is off,
            # so message age is NOT a valid physical RC connection test.
            # Use ArduPilot's receiver health / Radio Failsafe instead.
            with telemetry_state.lock:
                telemetry_connected = telemetry_state.connected
                rc_failsafe = telemetry_state.rc_failsafe
                rc_percent_available = (
                    telemetry_state.rc_rssi_percent is not None
                    and telemetry_state.rc_rssi_percent > 0
                )

                if not telemetry_connected:
                    rc_connected = False

                elif rc_failsafe is True:
                    rc_connected = False

                elif rc_failsafe is False:
                    rc_connected = True

                else:
                    # Startup fallback before SYS_STATUS establishes the
                    # receiver-health state. A valid receiver LQ value means
                    # the receiver has at least been seen.
                    rc_connected = rc_percent_available

            # Draw these AFTER the HUD so they cover the pitch ladder.
            draw_disconnect_messages(
                display_frame,
                camera_connected=camera_connected,
                telemetry_connected=telemetry_connected,
                rc_connected=rc_connected,
            )

            cv2.imshow(
                WINDOW_NAME,
                display_frame,
            )

            # Detect Ctrl + number using Windows key state directly.
            # This is more reliable than depending on OpenCV to encode
            # modifier combinations in waitKeyEx().
            key = cv2.waitKeyEx(20)

            ctrl_down = False
            key_0_down = False
            key_1_down = False
            key_2_down = False

            try:
                import ctypes

                user32 = ctypes.windll.user32

                VK_CONTROL = 0x11
                VK_0 = 0x30
                VK_1 = 0x31
                VK_2 = 0x32

                ctrl_down = bool(
                    user32.GetAsyncKeyState(VK_CONTROL) & 0x8000
                )
                key_0_down = bool(
                    user32.GetAsyncKeyState(VK_0) & 0x8000
                )
                key_1_down = bool(
                    user32.GetAsyncKeyState(VK_1) & 0x8000
                )
                key_2_down = bool(
                    user32.GetAsyncKeyState(VK_2) & 0x8000
                )

            except Exception:
                pass

            # Latch so holding the keys does not print every frame.
            if not hasattr(main, "_ctrl_combo_latched"):
                main._ctrl_combo_latched = False

            combo_active = (
                ctrl_down
                and (
                    key_0_down
                    or key_1_down
                    or key_2_down
                )
            )

            if combo_active and not main._ctrl_combo_latched:
                if key_0_down:
                    test_alert_mode = 0
                    print("Alert test mode: NORMAL (Ctrl+0)")

                elif key_1_down:
                    test_alert_mode = 1
                    print("Alert test mode: YELLOW WARNING (Ctrl+1)")

                elif key_2_down:
                    test_alert_mode = 2
                    print("Alert test mode: RED CRITICAL (Ctrl+2)")

                main._ctrl_combo_latched = True

            elif not combo_active:
                main._ctrl_combo_latched = False

            key_low = key & 0xFF

            # q / Esc still quit normally.
            if key_low == ord("q") or key_low == 27:
                break

            if (
                cv2.getWindowProperty(
                    WINDOW_NAME,
                    cv2.WND_PROP_VISIBLE,
                )
                < 1
            ):
                break

    except KeyboardInterrupt:
        print("\nProgram stopped by user.")

    finally:
        stop_event.set()

        if cap is not None:
            cap.release()

        telemetry_thread.join(timeout=2.0)
        cv2.destroyAllWindows()

    


if __name__ == "__main__":
    main()